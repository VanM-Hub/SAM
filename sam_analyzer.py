"""
SAM Analyzer — jalur analisis file NYATA (eksperimen vertical slice).

Ini BUKAN bagian governance/arsitektur SAM. Ini modul fungsional mandiri yang
membuktikan SAM mampu melakukan "Real Action": membaca file dari disk,
menganalisis isi, dan menghasilkan laporan + audit trail yang bisa diverifikasi.

Dirancang deterministik dan offline-friendly. Tidak menyentuh sistem selain
file yang ditunjuk user.

Penggunaan:
    python sam_analyzer.py <path_file> [--out <file_laporan>]

Contoh:
    python sam_analyzer.py "D:\\data\laporan.xlsx"
    python sam_analyzer.py "app.log" --out "hasil_analisis.txt"
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from typing import Any, Dict, List


# ---------------------------------------------------------------------------
# Audit trail (sederhana, kredibel, bisa diverifikasi)
# ---------------------------------------------------------------------------

class AuditTrail:
    """Merekam setiap aksi nyata yang dilakukan analyzer (jejak audit)."""

    def __init__(self) -> None:
        self._entries: List[Dict[str, Any]] = []

    def record(self, action: str, detail: str, extra: Dict[str, Any] | None = None) -> None:
        self._entries.append({
            "ts": datetime.now(timezone.utc).isoformat(),
            "action": action,
            "detail": detail,
            **(extra or {}),
        })

    @property
    def entries(self) -> List[Dict[str, Any]]:
        return list(self._entries)


# ---------------------------------------------------------------------------
# Analisis file Excel / CSV
# ---------------------------------------------------------------------------

def _read_tabular(path: str, audit: AuditTrail) -> Dict[str, Any]:
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xls"):
        import openpyxl
        audit.record("read_excel", path, {"ext": ext})
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            sheets = {}
            for ws in wb.worksheets:
                rows: List[List[Any]] = []
                for row in ws.iter_rows(values_only=True):
                    rows.append(list(row))
                sheets[ws.title] = {"dims": f"{ws.max_row}x{ws.max_column}", "rows": rows}
            return {"type": "excel", "sheets": sheets}
        finally:
            wb.close()
    else:  # csv / tsv
        sep = "\t" if ext == ".tsv" else ","
        audit.record("read_csv", path, {"ext": ext, "sep": sep})
        with open(path, "r", encoding="utf-8-sig", errors="replace") as fh:
            reader = csv.reader(fh, delimiter=sep)
            rows = [row for row in reader]
        return {"type": "csv", "sheets": {"Sheet1": {"dims": f"{len(rows)}x{max((len(r) for r in rows), default=0)}", "rows": rows}}}


def _analyze_tabular(data: Dict[str, Any], path: str, audit: AuditTrail) -> Dict[str, Any]:
    findings: List[Dict[str, Any]] = []
    summary: Dict[str, Any] = {}

    for sheet_name, sheet in data["sheets"].items():
        rows = sheet["rows"]
        total_rows = len(rows)
        if total_rows == 0:
            findings.append({"sheet": sheet_name, "type": "empty_sheet", "msg": "Sheet kosong (tidak ada baris)."})
            continue

        header = rows[0]
        n_cols = len(header)
        body = rows[1:] if total_rows > 1 else []

        issues = {"empty_cells": 0, "total_cells": 0, "mixed_type_cols": [], "dup_rows": 0}
        col_types: List[set] = [set() for _ in range(n_cols)]
        seen: set = set()

        for r_i, row in enumerate(body):
            # sel kosong
            for c_i in range(n_cols):
                cell = row[c_i] if c_i < len(row) else None
                val = "" if cell is None else cell
                issues["total_cells"] += 1
                if val == "" or val is None:
                    issues["empty_cells"] += 1
                else:
                    col_types[c_i].add(type(val).__name__)
            # baris duplikat
            key = tuple(row)
            if key in seen:
                issues["dup_rows"] += 1
            else:
                seen.add(key)

        # deteksi tipe campuran per kolom
        for c_i in range(n_cols):
            if len(col_types[c_i]) > 1:
                mixed = col_types[c_i]
                issues["mixed_type_cols"].append({
                    "col": c_i,
                    "col_name": str(header[c_i]) if c_i < len(header) else f"col{c_i}",
                    "types": sorted(mixed),
                })

        findings.append({
            "sheet": sheet_name,
            "type": "sheet_scan",
            "rows": total_rows,
            "data_rows": len(body),
            "cols": n_cols,
            "empty_cells": issues["empty_cells"],
            "total_cells": issues["total_cells"],
            "mixed_type_cols": issues["mixed_type_cols"],
            "duplicate_rows": issues["dup_rows"],
            "issues_count": issues["empty_cells"] + issues["dup_rows"] + len(issues["mixed_type_cols"]),
        })
        summary[sheet_name] = issues

    audit.record("analyze_tabular", path, {"sheets": list(summary.keys())})
    return {"findings": findings, "summary": summary, "total_issues": sum(f.get("issues_count", 0) for f in findings)}


# ---------------------------------------------------------------------------
# Analisis file log / teks
# ---------------------------------------------------------------------------

LOG_LEVELS = ("error", "warn", "warning", "critical", "fatal", "traceback", "exception", "fail", "timeout")
IP_RE = re.compile(r"\b\d{1,3}(?:\.\d{1,3}){3}\b")
EMAIL_RE = re.compile(r"\b[\w.+-]+@[\w-]+\.[\w.]+\b")
TS_RE = re.compile(r"\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}:\d{2}")


def _analyze_text(path: str, audit: AuditTrail) -> Dict[str, Any]:
    audit.record("read_text", path)
    with open(path, "r", encoding="utf-8", errors="replace") as fh:
        lines = fh.readlines()

    total = len(lines)
    empty = sum(1 for ln in lines if ln.strip() == "")
    long_lines = sum(1 for ln in lines if len(ln) > 500)

    # deteksi level penting
    level_matches = {lvl: 0 for lvl in LOG_LEVELS}
    keyword_lines: List[int] = []
    for i, ln in enumerate(lines, 1):
        lower = ln.lower()
        for lvl in LOG_LEVELS:
            if lvl in lower:
                level_matches[lvl] += 1
                keyword_lines.append(i)

    # pola berulang (5 baris paling umum)
    trimmed = (ln.strip() for ln in lines if ln.strip())
    repeats = Counter(trimmed).most_common(5)
    repeated_lines = [{"text": t[:120], "count": c} for t, c in repeats if c > 1]

    # timestamp range
    timestamps = [TS_RE.search(ln).group() for ln in lines if TS_RE.search(ln)]
    ts_range = None
    if timestamps:
        ts_range = {"first": min(timestamps), "last": max(timestamps), "count": len(timestamps)}

    ips = len(set(IP_RE.findall("".join(lines))))
    emails = len(set(EMAIL_RE.findall("".join(lines))))

    total_issues = sum(level_matches.values()) + len(repeated_lines)
    audit.record("analyze_text", path, {"lines": total})

    return {
        "findings": [
            {"type": "file_meta", "lines": total, "empty_lines": empty, "long_lines": long_lines},
            {"type": "log_levels", "counts": {k: v for k, v in level_matches.items() if v > 0}},
            {"type": "keyword_lines", "lines": keyword_lines[:50], "total": len(keyword_lines)},
            {"type": "repeated_lines", "patterns": repeated_lines},
            {"type": "ts_range", "range": ts_range},
            {"type": "entities", "unique_ips": ips, "unique_emails": emails},
        ],
        "total_issues": total_issues,
    }


def _analyze_file(path: str, audit: AuditTrail) -> Dict[str, Any]:
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xls", ".csv", ".tsv"):
        data = _read_tabular(path, audit)
        return _analyze_tabular(data, path, audit)
    else:
        return _analyze_text(path, audit)


# ---------------------------------------------------------------------------
# Laporan
# ---------------------------------------------------------------------------

def _render_report(path: str, result: Dict[str, Any], audit: AuditTrail, duration_ms: int) -> str:
    name = os.path.basename(path)
    ext = os.path.splitext(path)[1].lower()
    is_tabular = ext in (".xlsx", ".xlsm", ".xls", ".csv", ".tsv")
    title = "ANALISIS FILE EXCEL/CSV" if is_tabular else "ANALISIS FILE LOG/TEKS"

    lines = []
    lines.append("=" * 62)
    lines.append(f"  {title}")
    lines.append("=" * 62)
    lines.append(f"  Sumber       : {path}")
    lines.append(f"  Waktu        : {datetime.now(timezone.utc).isoformat()}")
    lines.append(f"  Durasi       : {duration_ms} ms")
    lines.append("")

    if is_tabular:
        for f in result["findings"]:
            if f["type"] == "empty_sheet":
                lines.append(f"[KOSONG] Sheet '{f['sheet']}': {f['msg']}")
            elif f["type"] == "sheet_scan":
                lines.append(f"--- Sheet '{f['sheet']}' ---")
                lines.append(f"  Baris total : {f['rows']} (data: {f['data_rows']})")
                lines.append(f"  Kolom       : {f['cols']}")
                lines.append(f"  Sel kosong  : {f['empty_cells']}/{f['total_cells']}")
                lines.append(f"  Baris duplikat: {f['duplicate_rows']}")
                if f["mixed_type_cols"]:
                    for m in f["mixed_type_cols"]:
                        lines.append(f"  Tipe campuran di kolom '{m['col_name']}' (idx {m['col']}): {', '.join(m['types'])}")
                else:
                    lines.append("  Tipe data   : konsisten di semua kolom")
        lines.append("")
        lines.append(f"TOTAL ISU DITEMUKAN: {result['total_issues']}")
    else:
        for f in result["findings"]:
            t = f["type"]
            if t == "file_meta":
                lines.append(f"  Baris total : {f['lines']}")
                lines.append(f"  Baris kosong: {f['empty_lines']}")
                lines.append(f"  Baris >500  : {f['long_lines']}")
            elif t == "log_levels":
                if f["counts"]:
                    lines.append("  Level penting terdeteksi:")
                    for k, v in f["counts"].items():
                        lines.append(f"    - {k}: {v} baris")
                else:
                    lines.append("  Level penting: tidak ada error/warning terdeteksi")
            elif t == "keyword_lines":
                lines.append(f"  Baris berisi kata kunci: {f['total']} (contoh baris: {f['lines'][:5]})")
            elif t == "repeated_lines":
                if f["patterns"]:
                    lines.append("  Pola baris berulang (kemungkinan masalah):")
                    for p in f["patterns"]:
                        lines.append(f"    - {p['count']}x: {p['text']}")
                else:
                    lines.append("  Tidak ada pola baris berulang")
            elif t == "ts_range":
                if f["range"]:
                    r = f["range"]
                    lines.append(f"  Rentang waktu : {r['first']} .. {r['last']} ({r['count']} timestamp)")
                else:
                    lines.append("  Timestamp     : tidak terdeteksi")
            elif t == "entities":
                lines.append(f"  IP unik   : {f['unique_ips']}")
                lines.append(f"  Email unik: {f['unique_emails']}")
        lines.append("")
        lines.append(f"TOTAL ISU DITEMUKAN: {result['total_issues']}")

    lines.append("")
    lines.append("--- JEJAK AUDIT (aksi nyata yang dilakukan) ---")
    for e in audit.entries:
        lines.append(f"  [{e['ts']}] {e['action']}: {e['detail']}")
    lines.append("")
    lines.append("=" * 62)
    return "\n".join(lines)


def main(argv: List[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="SAM Analyzer — analisis file real")
    parser.add_argument("path", help="Path file (Excel/CSV/log/teks)")
    parser.add_argument("--out", help="Simpan laporan ke file (jika diisi)")
    args = parser.parse_args(argv)

    if not os.path.isfile(args.path):
        print(f"ERROR: file tidak ditemukan: {args.path}", file=sys.stderr)
        return 2

    audit = AuditTrail()
    t0 = datetime.now()
    result = _analyze_file(args.path, audit)
    duration_ms = int((datetime.now() - t0).total_seconds() * 1000)
    report = _render_report(args.path, result, audit, duration_ms)

    print(report)

    if args.out:
        with open(args.out, "w", encoding="utf-8") as fh:
            fh.write(report)
        audit.record("write_report", args.out)
        print(f"\n[Laporan disimpan ke: {args.out}]")

    json_path = (args.out or args.path) + ".json"
    with open(json_path, "w", encoding="utf-8") as fh:
        json.dump({"source": args.path, "result": result, "audit": audit.entries}, fh, indent=2, default=str)
    audit.record("write_json", json_path)
    print(f"[Data JSON disimpan ke: {json_path}]")
    return 0


if __name__ == "__main__":
    sys.exit(main())
